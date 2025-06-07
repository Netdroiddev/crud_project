from django.shortcuts import render, redirect
from flask import json

import openpyxl

from django.shortcuts import render, redirect
 # Asegúrate de que el formulario esté bien importado
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect

from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
# views.py
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

import json

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
import json
from .forms import FormularioForm

@csrf_exempt
def registro_atencion_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            form = FormularioForm(data)
            
            if form.is_valid():
                form.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Formulario guardado correctamente'
                })
            else:
                errors = {field: error[0] for field, error in form.errors.items()}
                return JsonResponse({
                    'success': False,
                    'errors': errors
                }, status=400)
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error en el servidor: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método no permitido'
    }, status=405)


def exito(request):
    return render(request, 'crud_app/exito.html')


from django.shortcuts import render



def inicio(request):
    return render(request, 'crud_app/base.html')


def index(request):
    return render(request, 'crud_app/index.html')

def export_credentials_to_excel(request):
    """Genera y descarga un archivo Excel con las credenciales"""
    credentials = Credential.objects.all().order_by('-created_at')
    
    # Crear el libro de trabajo y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credenciales"
    
    # Definir los encabezados y aplicar estilos
    headers = ["ID", "Nombre", "Usuario", "Gmail", "Estado", "Creado"]
    ws.append(headers)
    
    # Estilos para los encabezados
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20  # Ancho en píxeles aprox.
    
    # Agregar los datos
    for cred in credentials:
        ws.append([
            cred.id,
            cred.nombre,
            cred.usuario,
            cred.gmail,
            cred.estado,
            cred.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Preparar la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="credenciales.xlsx"'
    wb.save(response)
    return response
